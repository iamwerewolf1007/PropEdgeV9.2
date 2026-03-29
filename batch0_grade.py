#!/usr/bin/env python3
"""
PropEdge V9.2 — BATCH 0: GRADE + UPDATE (6:00 AM UK)
=========================================================
Runs after games finish. Does in order:
1. Fetch box scores from nba_api (ScoreboardV3 + BoxScoreTraditionalV3)
2. Append new game rows to nba_gamelogs_2025_26.csv
   - Played rows: full stats + live rolling computed
   - DNP stubs: DNP=1, MIN_NUM=0, all stats NaN, rolling all NaN
3. Grade plays in season_2025_26.json and today.json
   - Result=WIN/LOSS based on actual PTS vs line
   - Result=DNP if player not in box scores
   - Graded plays are IMMUTABLE once set
4. Generate post-match reasoning (via reasoning_engine) using actual box score data
5. Cross-check: recompute rolling stats fresh vs stored in play — flag deviations
6. Update daily Excel Graded sheet
7. Rebuild H2H database
8. Retrain projection model
9. Git push
"""
import pandas as pd
import numpy as np
import json, sys, time, re, subprocess
from pathlib import Path
from datetime import datetime, timedelta

sys.path.insert(0, str(Path(__file__).parent.resolve()))
from config import *
from audit import log_event, log_file_state, verify_no_deletion, log_batch_summary
from rolling_engine import filter_played, compute_rolling_for_new_rows, \
                           build_player_index, get_prior_games_played, extract_prediction_features
from reasoning_engine import generate_post_match_reason


def _clean(obj):
    return clean_json(obj)


def notify(title, msg):
    try:
        subprocess.run(['osascript','-e',f'display notification "{msg}" with title "{title}"'],
                       capture_output=True, timeout=5)
    except: pass


def git_push(repo, message):
    env = {**__import__('os').environ,
           'GIT_SSH_COMMAND': 'ssh -o BatchMode=yes -o StrictHostKeyChecking=no'}
    try:
        subprocess.run(['git','remote','set-url','origin',GIT_REMOTE],
                       cwd=repo, capture_output=True, timeout=10, env=env)
        subprocess.run(['git','add','-A'], cwd=repo, capture_output=True, timeout=30, env=env)
        r = subprocess.run(['git','commit','-m',message],
                           cwd=repo, capture_output=True, timeout=30, env=env)
        if r.returncode != 0:
            if 'nothing to commit' in (r.stderr.decode() + r.stdout.decode()):
                print("  ✓ Git: nothing to commit"); return
        r3 = subprocess.run(['git','push'], cwd=repo, capture_output=True, timeout=60, env=env)
        if r3.returncode != 0:
            err = r3.stderr.decode()
            if 'set-upstream' in err or 'no upstream' in err:
                r3 = subprocess.run(['git','push','--set-upstream','origin','main'],
                                    cwd=repo, capture_output=True, timeout=60, env=env)
        print(f"  {'✓' if r3.returncode==0 else '⚠'} Git push: {message}")
        if r3.returncode != 0:
            print(f"    {r3.stderr.decode()[:200]}")
    except subprocess.TimeoutExpired:
        print("  ⚠ Git push timed out")
    except Exception as e:
        print(f"  ⚠ Git error: {e}")


def _si(v):
    try: return int(v) if pd.notna(v) else 0
    except: return 0


def _parse_min(v):
    s = str(v).strip()
    if s in ('','None','nan','0','PT00M00.00S'): return 0.0
    if s.startswith('PT') and 'M' in s:
        m = re.match(r'PT(\d+)M([\d.]+)S', s)
        return float(m.group(1)) + float(m.group(2))/60 if m else 0.0
    if ':' in s:
        p = s.split(':'); return float(p[0]) + float(p[1])/60
    try: return float(s)
    except: return 0.0


# ─── FETCH BOX SCORES ────────────────────────────────────────
def fetch_boxscores(date_str):
    """Fetch NBA box scores. Returns (played_rows, players_in_box_set)."""
    from nba_api.stats.endpoints import ScoreboardV3, BoxScoreTraditionalV3
    print(f"\n  Fetching box scores: {date_str}...")
    time.sleep(1)

    sb = ScoreboardV3(game_date=date_str, league_id='00')
    gh = sb.game_header.get_data_frame()
    ls = sb.line_score.get_data_frame()
    if gh.empty:
        print("    No games found"); return [], set()

    gids = gh['gameId'].tolist()
    print(f"    {len(gids)} games")

    ctx = {}
    for g in gids:
        r = ls[ls['gameId'] == g]
        if len(r) >= 2:
            ctx[str(g)] = {
                'htid': r.iloc[0]['teamId'],
                'ht':   r.iloc[0]['teamTricode'],
                'at':   r.iloc[1]['teamTricode'],
                'hs':   _si(r.iloc[0].get('score', 0)),
                'as_':  _si(r.iloc[1].get('score', 0)),
            }

    # Bio cache from existing game log
    df26 = pd.read_csv(FILE_GL_2526)
    bio  = {}
    bc   = ['PLAYER_ID','PLAYER_NAME','PLAYER_POSITION','PLAYER_POSITION_FULL',
            'PLAYER_CURRENT_TEAM','GAME_TEAM_ABBREVIATION','GAME_TEAM_NAME',
            'PLAYER_HEIGHT','PLAYER_WEIGHT','PLAYER_EXPERIENCE','PLAYER_COUNTRY',
            'PLAYER_DRAFT_YEAR','PLAYER_DRAFT_ROUND','PLAYER_DRAFT_NUMBER']
    for _, r in df26.drop_duplicates('PLAYER_ID', keep='last')[bc].iterrows():
        bio[r['PLAYER_ID']] = r.to_dict()

    rows = []
    players_in_box = set()  # All player names seen in box scores (incl. DNP in game)

    for g in gids:
        time.sleep(0.8)
        try:
            box = BoxScoreTraditionalV3(game_id=g)
            ps  = box.player_stats.get_data_frame()
            if ps.empty: continue

            col_map = {
                'personId':'PLAYER_ID','teamId':'TEAM_ID','teamTricode':'TEAM_ABBREVIATION',
                'firstName':'FN','familyName':'LN','minutes':'MR',
                'fieldGoalsMade':'FGM','fieldGoalsAttempted':'FGA',
                'threePointersMade':'FG3M','threePointersAttempted':'FG3A',
                'freeThrowsMade':'FTM','freeThrowsAttempted':'FTA',
                'reboundsOffensive':'OREB','reboundsDefensive':'DREB','reboundsTotal':'REB',
                'assists':'AST','steals':'STL','blocks':'BLK','turnovers':'TOV',
                'foulsPersonal':'PF','points':'PTS','plusMinusPoints':'PLUS_MINUS',
            }
            ps = ps.rename(columns={k:v for k,v in col_map.items() if k in ps.columns})
            if 'PLAYER_NAME' not in ps.columns and 'FN' in ps.columns:
                ps['PLAYER_NAME'] = ps['FN'].fillna('') + ' ' + ps['LN'].fillna('')

            c = ctx.get(str(g), {})

            for _, p in ps.iterrows():
                pname = str(p.get('PLAYER_NAME','')).strip()
                if pname: players_in_box.add(pname)

                mn = _parse_min(p.get('MR', 0))
                if mn <= 0:
                    continue  # DNP in game — not written as played row

                pid = _si(p.get('PLAYER_ID', 0))
                tid = _si(p.get('TEAM_ID', 0))
                ta  = str(p.get('TEAM_ABBREVIATION',''))
                ih  = 1 if tid == c.get('htid') else 0
                opp_rows = ps[ps['TEAM_ID'] != tid]['TEAM_ABBREVIATION']
                opp = opp_rows.iloc[0] if len(opp_rows) > 0 else 'UNK'
                mu  = f"{ta} vs. {opp}" if ih else f"{ta} @ {opp}"
                wl  = ('W' if c.get('hs',0) > c.get('as_',0) else 'L') if ih \
                       else ('W' if c.get('as_',0) > c.get('hs',0) else 'L')

                pts  = _si(p.get('PTS',0));  fgm = _si(p.get('FGM',0)); fga = _si(p.get('FGA',0))
                fg3m = _si(p.get('FG3M',0)); fg3a= _si(p.get('FG3A',0))
                ftm  = _si(p.get('FTM',0));  fta = _si(p.get('FTA',0))
                oreb = _si(p.get('OREB',0)); dreb= _si(p.get('DREB',0)); reb = _si(p.get('REB',0))
                ast  = _si(p.get('AST',0));  stl = _si(p.get('STL',0));  blk = _si(p.get('BLK',0))
                tov  = _si(p.get('TOV',0));  pf  = _si(p.get('PF',0));   pm  = _si(p.get('PLUS_MINUS',0))

                fgp = fgm/fga if fga>0 else 0.0
                f3p = fg3m/fg3a if fg3a>0 else 0.0
                ftp = ftm/fta  if fta>0 else 0.0
                efg = (fgm+0.5*fg3m)/fga if fga>0 else 0.0
                tsa = 2*(fga+0.44*fta)
                ts  = pts/tsa if tsa>0 else 0.0
                usg = (fga+0.44*fta+tov)/(mn/5) if mn>0 else 0.0
                pra = pts+reb+ast
                ddc = sum(1 for x in [pts,reb,ast,stl,blk] if x>=10)
                dd  = 1 if ddc>=2 else 0; td = 1 if ddc>=3 else 0
                fp  = pts+1.25*reb+1.5*ast+2*stl+2*blk-0.5*tov+0.5*fg3m+1.5*dd+3*td
                b   = bio.get(pid, {})

                rows.append({
                    'PLAYER_ID':pid,
                    'PLAYER_NAME': pname or b.get('PLAYER_NAME',''),
                    'SEASON':'2025-26','SEASON_TYPE':'Regular Season',
                    'PLAYER_POSITION':b.get('PLAYER_POSITION',''),
                    'PLAYER_POSITION_FULL':b.get('PLAYER_POSITION_FULL',''),
                    'PLAYER_CURRENT_TEAM':b.get('PLAYER_CURRENT_TEAM',ta),
                    'GAME_TEAM_ABBREVIATION':ta,'GAME_TEAM_NAME':b.get('GAME_TEAM_NAME',''),
                    'PLAYER_HEIGHT':b.get('PLAYER_HEIGHT',''),
                    'PLAYER_WEIGHT':b.get('PLAYER_WEIGHT',0),
                    'PLAYER_EXPERIENCE':b.get('PLAYER_EXPERIENCE',0),
                    'PLAYER_COUNTRY':b.get('PLAYER_COUNTRY',''),
                    'PLAYER_DRAFT_YEAR':b.get('PLAYER_DRAFT_YEAR',0),
                    'PLAYER_DRAFT_ROUND':b.get('PLAYER_DRAFT_ROUND',0),
                    'PLAYER_DRAFT_NUMBER':b.get('PLAYER_DRAFT_NUMBER',0),
                    'GAME_ID':int(g),'GAME_DATE':date_str,'MATCHUP':mu,'OPPONENT':opp,
                    'IS_HOME':ih,'WL':wl,'WL_WIN':1 if wl=='W' else 0,
                    'WL_LOSS':1 if wl=='L' else 0,'GAMES_PLAYED_SEASON_RUNNING':0,
                    'MIN':int(round(mn)),'MIN_NUM':round(mn,1),
                    'FGM':fgm,'FGA':fga,'FG_PCT':round(fgp,4),
                    'FG3M':fg3m,'FG3A':fg3a,'FG3_PCT':round(f3p,4),
                    'FTM':ftm,'FTA':fta,'FT_PCT':round(ftp,4),
                    'OREB':oreb,'DREB':dreb,'REB':reb,'AST':ast,
                    'STL':stl,'BLK':blk,'TOV':tov,'PF':pf,
                    'PTS':pts,'PLUS_MINUS':pm,'VIDEO_AVAILABLE':1,
                    'EFF_FG_PCT':round(efg,4),'TRUE_SHOOTING_PCT':round(ts,4),
                    'USAGE_APPROX':round(usg,2),'PTS_REB_AST':pra,
                    'PTS_REB':pts+reb,'PTS_AST':pts+ast,'REB_AST':reb+ast,
                    'DOUBLE_DOUBLE':dd,'TRIPLE_DOUBLE':td,'FANTASY_PTS':round(fp,2),
                    'SEASON_ID':22025,'DNP':0,
                })
        except Exception as e:
            print(f"    ✗ game {g}: {e}")

    print(f"  Fetched {len(rows)} played rows, {len(players_in_box)} players in box")
    log_event('B0','BOXSCORES_FETCHED',detail=f'{len(rows)} rows for {date_str}')
    return rows, players_in_box


# ─── APPEND GAME LOGS ─────────────────────────────────────────
def append_gamelogs(played_rows, dnp_player_names, date_str):
    """
    Append played rows + DNP stubs to nba_gamelogs_2025_26.csv.
    DNP stubs: DNP=1, MIN_NUM=0, all stats NaN, all rolling NaN.
    Played rows: rolling computed live (excluding DNP rows from windows).
    """
    df25 = pd.read_csv(FILE_GL_2425, parse_dates=['GAME_DATE'])
    df26 = pd.read_csv(FILE_GL_2526, parse_dates=['GAME_DATE'])
    rows_before = len(df26)
    log_file_state('B0', FILE_GL_2526, 'BEFORE_APPEND')

    # Ensure DNP col in df26
    if 'DNP' not in df26.columns: df26['DNP'] = 0
    df26['DNP'] = df26['DNP'].fillna(0).astype(int)

    # Build DNP stubs
    dnp_stubs = []
    for pname in dnp_player_names:
        stub = {c: np.nan for c in df26.columns}
        stub.update({'PLAYER_NAME':pname,'GAME_DATE':date_str,
                     'DNP':1,'MIN_NUM':0,'PTS':np.nan,'SEASON':'2025-26',
                     'SEASON_TYPE':'Regular Season','SEASON_ID':22025})
        dnp_stubs.append(stub)

    # Combine played + stubs
    all_new = played_rows + dnp_stubs
    if not all_new:
        print("  No new rows to append"); return

    ndf = pd.DataFrame(all_new)
    ndf['GAME_DATE'] = pd.to_datetime(ndf['GAME_DATE'])
    if 'DNP' not in ndf.columns: ndf['DNP'] = 0
    ndf['DNP'] = ndf['DNP'].fillna(0).astype(int)

    # Compute rolling for new rows (DNPs get NaN rolling, played get live rolling)
    hist = pd.concat([df25, df26], ignore_index=True)
    hist['GAME_DATE'] = pd.to_datetime(hist['GAME_DATE'])
    if 'DNP' not in hist.columns: hist['DNP'] = 0

    ndf = compute_rolling_for_new_rows(ndf, hist)

    # Align columns to df26 schema
    if 'DNP' not in df26.columns: df26['DNP'] = 0
    for c in df26.columns:
        if c not in ndf.columns: ndf[c] = np.nan
    ndf = ndf[[c for c in df26.columns if c in ndf.columns]]
    # Add any new cols not in df26 (e.g. DNP first time)
    for c in ndf.columns:
        if c not in df26.columns: df26[c] = np.nan

    updated = pd.concat([df26, ndf], ignore_index=True)
    updated = updated.sort_values(['PLAYER_NAME','GAME_DATE']).reset_index(drop=True)
    before_dedup = len(updated)
    updated = updated.drop_duplicates(subset=['PLAYER_NAME','GAME_DATE'], keep='last')
    if len(updated) < before_dedup:
        print(f"  Dedup: removed {before_dedup-len(updated)} duplicate rows")

    updated.to_csv(FILE_GL_2526, index=False)
    verify_no_deletion('B0', FILE_GL_2526, rows_before, len(updated), 'APPEND_GAMELOGS')
    print(f"  ✓ Game logs: {rows_before} → {len(updated)} "
          f"(+{len(played_rows)} played, +{len(dnp_stubs)} DNP stubs)")


# ─── CROSS-CHECK ROLLING STATS ────────────────────────────────
def crosscheck_rolling_stats(plays_for_date, date_str):
    """
    Recompute rolling stats fresh from game log CSV and compare against
    stored values in the play records. Flags deviations > 1pt on L30.
    Returns dict: player_name → integrity_flag_str (or None if OK)
    """
    print(f"\n  Cross-checking rolling stats for {date_str}...")
    integrity = {}

    try:
        from rolling_engine import load_combined, build_player_index
        combined = load_combined(FILE_GL_2425, FILE_GL_2526)
        pidx     = build_player_index(combined)

        flagged = 0
        for p in plays_for_date:
            player     = p.get('player','')
            line       = float(p.get('line', 0) or 0)
            stored_l30 = p.get('l30')

            if stored_l30 is None:
                integrity[player] = None
                continue

            prior_played = get_prior_games_played(pidx, player, date_str)
            if len(prior_played) < 5:
                integrity[player] = None
                continue

            feats = extract_prediction_features(prior_played, line)
            if feats is None:
                integrity[player] = None
                continue

            fresh_l30 = feats['L30']
            diff      = abs(fresh_l30 - float(stored_l30))

            if diff > 1.0:
                flag = (f"L30 stored={stored_l30:.1f} recomputed={fresh_l30:.1f} "
                        f"deviation={diff:.1f}pts — possible data drift")
                integrity[player] = flag
                log_event('B0','ROLLING_CROSSCHECK_FAIL',detail=f"{player}: {flag}")
                flagged += 1
            else:
                integrity[player] = None

        if flagged:
            print(f"  ⚠ Cross-check: {flagged} rolling stat deviations (>1pt) — see audit_log.csv")
        else:
            print(f"  ✓ Cross-check: all rolling stats consistent for {date_str}")

    except Exception as e:
        print(f"  ⚠ Cross-check failed: {e}")

    return integrity


# ─── GRADE PLAYS ─────────────────────────────────────────────
def grade_plays(date_str, played_rows, players_in_box):
    """
    Grade plays in season_2025_26.json and today.json.
    - WIN/LOSS based on actual PTS vs line
    - DNP if player not in box scores (play preserved, not dropped)
    - Generates post-match reasoning with actual box score data
    - Graded plays are IMMUTABLE (won't be overwritten)
    Returns: (dnp_player_names, plays_for_crosscheck)
    """
    # Build results map: player_name → row dict (with PTS, MIN_NUM, FGM, FGA, FG_PCT)
    results_map = {}
    for r in played_rows:
        pname = r.get('PLAYER_NAME','')
        if pname:
            results_map[pname] = r

    dnp_names         = []
    plays_for_check   = []
    graded = wins = losses = dnps = 0
    now_str = now_uk().strftime('%Y-%m-%d %H:%M')

    for fpath in [SEASON_2526, TODAY_JSON]:
        if not fpath.exists(): continue
        with open(fpath) as f: plays = json.load(f)
        changed = False

        for p in plays:
            if p.get('date') != date_str: continue
            if p.get('result') in ('WIN','LOSS'): continue  # IMMUTABLE

            player = p.get('player','')
            plays_for_check.append(p)

            # Player not in box at all → DNP
            if player not in players_in_box and player not in results_map:
                p['result']    = 'DNP'
                p['actualPts'] = None
                dnps += 1; changed = True
                if fpath == TODAY_JSON:
                    dnp_names.append(player)
                continue

            # Player in box but played 0 min → DNP
            box_row = results_map.get(player)
            if box_row is None:
                p['result']    = 'DNP'
                p['actualPts'] = None
                dnps += 1; changed = True
                if fpath == TODAY_JSON:
                    dnp_names.append(player)
                continue

            actual = int(box_row.get('PTS', 0))
            p['actualPts'] = actual
            p['delta']     = round(actual - p['line'], 1)

            d = p.get('dir','')
            if 'OVER'  in d: p['result'] = 'WIN' if actual > p['line'] else 'LOSS'
            elif 'UNDER' in d: p['result'] = 'WIN' if actual < p['line'] else 'LOSS'
            else:
                p['result'] = 'DNP'; dnps += 1; changed = True; continue

            if p['result'] == 'WIN': wins += 1
            elif p['result'] == 'LOSS': losses += 1
            graded += 1; changed = True

            # Build actual_box dict for reasoning engine
            actual_min     = float(box_row.get('MIN_NUM', 0) or 0)
            actual_fgm     = int(box_row.get('FGM', 0) or 0)
            actual_fga     = int(box_row.get('FGA', 0) or 0)
            raw_fg         = box_row.get('FG_PCT', 0) or 0
            actual_fg_pct  = round(float(raw_fg) * 100, 1) if float(raw_fg) < 1.5 else round(float(raw_fg), 1)

            actual_box_data = {
                'actual_pts':    actual,
                'actual_min':    actual_min,
                'actual_fgm':    actual_fgm,
                'actual_fga':    actual_fga,
                'actual_fg_pct': actual_fg_pct,
                'integrity_flag': None,  # filled in after crosscheck
            }

            try:
                reason, loss_type = generate_post_match_reason(p, actual_box_data)
                p['postMatchReason'] = reason
                p['lossType']        = loss_type
            except Exception as e:
                p['postMatchReason'] = f"Graded {now_str}"
                p['lossType']        = None

        if changed:
            with open(fpath,'w') as f: json.dump(_clean(plays), f)

    print(f"  ✓ Graded {graded}: {wins}W/{losses}L, {dnps} DNP")
    log_batch_summary('B0', plays_graded=graded, wins=wins, losses=losses, dnp=dnps)
    return list(set(dnp_names)), plays_for_check


# ─── UPDATE POST-MATCH WITH INTEGRITY FLAGS ──────────────────
def apply_integrity_flags(date_str, integrity_map):
    """Append integrity flag to postMatchReason for any flagged players."""
    if not any(integrity_map.values()):
        return
    for fpath in [SEASON_2526, TODAY_JSON]:
        if not fpath.exists(): continue
        with open(fpath) as f: plays = json.load(f)
        changed = False
        for p in plays:
            if p.get('date') != date_str: continue
            flag = integrity_map.get(p.get('player',''))
            if flag and p.get('postMatchReason'):
                p['postMatchReason'] += f' ⚠ Data integrity: {flag}'
                changed = True
        if changed:
            with open(fpath,'w') as f: json.dump(_clean(plays), f)


# ─── UPDATE DAILY EXCEL ───────────────────────────────────────
def update_daily_excel(date_str, plays_for_date):
    """Fill in the Graded sheet of the daily Excel workbook."""
    excel_path = DAILY_DIR / f"{date_str}.xlsx"
    if not excel_path.exists(): return
    try:
        import openpyxl
        from openpyxl.styles import PatternFill
        wb = openpyxl.load_workbook(excel_path)
        if 'Graded' not in wb.sheetnames: return
        ws = wb['Graded']
        now_str = now_uk().strftime('%Y-%m-%d %H:%M')
        fills   = {'WIN': PatternFill('solid',fgColor='22C55E'),
                   'LOSS':PatternFill('solid',fgColor='EF4444'),
                   'DNP': PatternFill('solid',fgColor='F59E0B')}
        play_map = {p.get('player',''): p for p in plays_for_date}
        for row in ws.iter_rows(min_row=2):
            player = row[0].value
            if not player: continue
            p = play_map.get(player)
            if not p: continue
            row[4].value = p.get('actualPts')
            row[5].value = p.get('result','')
            row[6].value = p.get('delta')
            row[7].value = p.get('postMatchReason','')
            row[8].value = p.get('lossType','')
            row[9].value = now_str
            fill = fills.get(p.get('result',''))
            if fill:
                for cell in row: cell.fill = fill
        wb.save(excel_path)
        print(f"  ✓ Daily Excel graded: {excel_path.name}")
    except Exception as e:
        print(f"  ⚠ Daily Excel update failed: {e}")


# ─── MAIN ────────────────────────────────────────────────────
def main():
    print("="*60)
    print(f"PropEdge V9.2 — BATCH 0: GRADE + UPDATE")
    print(f"  {now_uk().strftime('%Y-%m-%d %H:%M %Z')}")
    print("="*60)
    log_event('B0','BATCH_START')

    yesterday = (datetime.now(get_et()) - timedelta(days=1)).strftime('%Y-%m-%d')
    print(f"  Grading: {yesterday}")

    # 1. Fetch box scores
    played_rows, players_in_box = fetch_boxscores(yesterday)

    # 2. Grade plays (before appending so we have results map)
    dnp_names, plays_for_check = grade_plays(yesterday, played_rows, players_in_box)

    # 3. Append game logs (played + DNP stubs)
    append_gamelogs(played_rows, dnp_names, yesterday)

    # 4. Cross-check rolling stats (now that new rows are in CSV)
    integrity_map = crosscheck_rolling_stats(plays_for_check, yesterday)

    # 5. Apply integrity flags to post-match reasoning
    apply_integrity_flags(yesterday, integrity_map)

    # 6. Update daily Excel
    update_daily_excel(yesterday, plays_for_check)

    # 7. Rebuild H2H
    print("\n  Rebuilding H2H...")
    from h2h_builder import build_h2h
    build_h2h(FILE_GL_2425, FILE_GL_2526, FILE_H2H)

    # 8. Retrain model
    print("  Retraining model...")
    from model_trainer import train_and_save
    train_and_save(FILE_GL_2425, FILE_GL_2526, FILE_H2H, FILE_MODEL, FILE_TRUST)

    # 9. Push
    repo = REPO_DIR if REPO_DIR.exists() else ROOT
    git_push(repo, f"B0: grade {yesterday}")
    log_event('B0','BATCH_COMPLETE')
    notify("PropEdge V9.2", f"B0 done — graded {yesterday}")
    print("  ✓ BATCH 0 complete")


if __name__ == '__main__': main()
