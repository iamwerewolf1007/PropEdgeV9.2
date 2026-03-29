#!/usr/bin/env python3
"""
PropEdge V9.2 — BATCH 1/2/3: PREDICT
Usage: python3 batch_predict.py [1|2|3] [YYYY-MM-DD]
"""
import pandas as pd
import numpy as np
import json, sys, time, pickle, requests, unicodedata, re
from pathlib import Path
from datetime import datetime, timedelta

sys.path.insert(0, str(Path(__file__).parent.resolve()))
from config import *
from audit import log_event, log_file_state, verify_no_deletion, log_batch_summary
from rolling_engine import load_combined, build_player_index, get_prior_games_played, extract_prediction_features
from reasoning_engine import generate_pre_match_reason

BATCH = int(sys.argv[1]) if len(sys.argv) > 1 and sys.argv[1] in ('1','2','3') else 2


# ─── NAME RESOLVER ───────────────────────────────────────────
_NICKNAMES = {
    'nic':'nicolas','nick':'nicolas','herb':'herbert','moe':'mohamed','cam':'cameron',
    'drew':'andrew','alex':'alexander','will':'william','kenny':'kenyon','mo':'mohamed',
    'greg':'gregory','matt':'matthew','mike':'michael','chris':'christopher',
    'jon':'jonathan','joe':'joseph','ben':'benjamin','dan':'daniel','dave':'david',
    'rob':'robert','bob':'robert','ed':'edward','jeff':'jeffrey','jake':'jacob',
    'tony':'anthony','tj':'tj','cj':'cj','pj':'pj','aj':'aj',
}

def _norm(n):
    n = unicodedata.normalize('NFKD', str(n)).encode('ascii','ignore').decode()
    n = n.replace('.','').replace("'",'').strip()
    n = re.sub(r'\s+', ' ', n)
    n = re.sub(r'\s+(Jr|Sr|II|III|IV|V)\s*$', '', n, flags=re.IGNORECASE)
    return n.lower().strip()

def build_name_map(pidx):
    return {_norm(k): k for k in pidx}

def resolve_name(odds_name, pidx, name_map):
    if odds_name in pidx: return odds_name
    n = _norm(odds_name)
    if n in name_map: return name_map[n]
    for sfx in ['jr','sr','ii','iii','iv']:
        if n + ' ' + sfx in name_map: return name_map[n + ' ' + sfx]
    parts = odds_name.strip().split()
    if len(parts) >= 2:
        first = parts[0].lower()
        if first in _NICKNAMES:
            expanded = _NICKNAMES[first] + ' ' + ' '.join(parts[1:])
            en = _norm(expanded)
            if en in name_map: return name_map[en]
            for sfx in ['jr','sr','ii','iii','iv']:
                if en + ' ' + sfx in name_map: return name_map[en + ' ' + sfx]
    return None


# ─── ODDS API ────────────────────────────────────────────────
def _check_credits(headers, label=''):
    r = headers.get('x-requests-remaining','?')
    print(f"    Credits: {r} {label}")
    if r != '?' and int(r) <= CREDIT_ALERT:
        print("    ⚠ LOW CREDITS")

def fetch_props(date_str):
    print(f"\n  Fetching props for {date_str} (Batch {BATCH})...")
    d  = datetime.strptime(date_str, '%Y-%m-%d')
    fr = (d - timedelta(hours=6)).strftime('%Y-%m-%dT%H:%M:%SZ')
    to = (d + timedelta(hours=30)).strftime('%Y-%m-%dT%H:%M:%SZ')

    r1 = requests.get(f"{ODDS_API_BASE}/sports/{SPORT}/events",
                      params={'apiKey':ODDS_API_KEY,'dateFormat':'iso',
                              'commenceTimeFrom':fr,'commenceTimeTo':to},
                      timeout=30)
    r1.raise_for_status(); _check_credits(r1.headers, 'events')

    et = get_et()
    events = [e for e in r1.json()
              if datetime.fromisoformat(e['commence_time'].replace('Z','+00:00'))
                         .astimezone(et).strftime('%Y-%m-%d') == date_str]
    print(f"    {len(events)} games on {date_str}")
    if not events: return {}, []

    games = {}; spreads = []
    for e in events:
        eid = e['id']; hr = e['home_team']; ar = e['away_team']
        ts  = e['commence_time']
        try:
            gt = datetime.fromisoformat(ts.replace('Z','+00:00')).astimezone(et)\
                         .strftime('%-I:%M %p ET')
        except: gt = ''
        ht = resolve_abr(hr); at = resolve_abr(ar)
        games[eid] = {'home':ht,'away':at,'home_raw':hr,'away_raw':ar,
                      'gt':gt,'ts':ts,'spread':None,'total':None,'props':{}}
        try:
            r2 = requests.get(f"{ODDS_API_BASE}/sports/{SPORT}/events/{eid}/odds",
                              params={'apiKey':ODDS_API_KEY,'regions':'us',
                                      'markets':'player_points,spreads,totals',
                                      'oddsFormat':'american','dateFormat':'iso'},
                              timeout=30)
            r2.raise_for_status(); _check_credits(r2.headers)
            d2 = r2.json(); g = games[eid]
            for bm in d2.get('bookmakers',[]):
                for m in bm.get('markets',[]):
                    mk = m.get('key','')
                    if mk == 'spreads' and g['spread'] is None:
                        for o in m.get('outcomes',[]):
                            if o.get('name') == hr: g['spread'] = o.get('point')
                    elif mk == 'totals' and g['total'] is None:
                        for o in m.get('outcomes',[]):
                            if o.get('name','').upper() == 'OVER': g['total'] = o.get('point')
                    elif mk == 'player_points':
                        for o in m.get('outcomes',[]):
                            pl = (o.get('description') or '').strip() or o.get('name','').strip()
                            pt = o.get('point'); sd = o.get('name','').upper(); pr = o.get('price')
                            if not pl or pt is None: continue
                            if pl not in g['props']:
                                g['props'][pl] = {'line':pt,'over':None,'under':None,'books':0}
                            if sd == 'OVER':  g['props'][pl]['over'] = pr;  g['props'][pl]['books'] += 1
                            elif sd == 'UNDER': g['props'][pl]['under'] = pr
            if g['spread'] is not None:
                spreads.append({'Date':date_str,'Game':f"{at} @ {ht}",'Home':ht,'Away':at,
                                'Spread':g['spread'],'Total':g['total'],'Commence':ts})
            print(f"    ✓ {at} @ {ht}: {len(g['props'])} props")
            time.sleep(0.3)
        except Exception as ex:
            print(f"    ✗ {ar} @ {hr}: {ex}"); time.sleep(1)

    tp = sum(len(g['props']) for g in games.values())
    print(f"  Total: {tp} props across {len(games)} games")
    log_event(f'B{BATCH}','PROPS_FETCHED',detail=f'{tp} props, {len(games)} games')
    return games, spreads


# ─── PREDICTION ENGINE ────────────────────────────────────────
def run_predictions(games, date_str):
    print(f"\n  Running predictions (live rolling, DNP-safe)...")

    combined = load_combined(FILE_GL_2425, FILE_GL_2526)
    h2h      = pd.read_csv(FILE_H2H)
    h2h_dedup = h2h.drop_duplicates(subset=['PLAYER_NAME','OPPONENT'], keep='last')
    h2h_lkp  = {(r['PLAYER_NAME'],r['OPPONENT']): r.to_dict()
                for _,r in h2h_dedup.iterrows()}
    pidx     = build_player_index(combined)
    name_map = build_name_map(pidx)

    # Pace rank from played games only
    from rolling_engine import filter_played
    played_only = filter_played(combined)
    team_fga  = played_only.groupby('OPPONENT')['FGA'].mean()
    pace_rank = {t: i+1 for i,(t,_) in enumerate(team_fga.sort_values(ascending=False).items())}

    # B2B map (played games only)
    b2b = {}
    for pn, g in played_only.sort_values('GAME_DATE').groupby('PLAYER_NAME'):
        ds = g['GAME_DATE'].values
        for i in range(len(ds)):
            k = (pn, pd.Timestamp(ds[i]).strftime('%Y-%m-%d'))
            b2b[k] = int((ds[i]-ds[i-1]).astype('timedelta64[D]').astype(int)) if i>0 else 99

    model = None
    if FILE_MODEL.exists():
        with open(FILE_MODEL,'rb') as f: model = pickle.load(f)
    trust = {}
    if FILE_TRUST.exists():
        with open(FILE_TRUST) as f: trust = json.load(f)

    # Existing plays for line history
    existing = []
    if TODAY_JSON.exists():
        with open(TODAY_JSON) as f: existing = json.load(f)
    exist_map = {(p['player'],p.get('match','')): p
                 for p in existing if p.get('date') == date_str}

    batch_ts = now_uk().strftime('%H:%M')
    plays    = []
    skipped  = {'low_line':0,'no_player':0,'few_games':0,'no_features':0}

    for eid, g in games.items():
        ht = g['home']; at = g['away']
        ms  = f"{at} @ {ht}"
        fms = f"{TEAM_FULL.get(at,at)} @ {TEAM_FULL.get(ht,ht)}"
        sv  = g['spread']; tv = g['total']
        blow = abs(sv) >= 10 if sv else False

        for pname_raw, pd_ in g['props'].items():
            line = pd_.get('line')
            if not line or line < 3: skipped['low_line'] += 1; continue

            pname = resolve_name(pname_raw, pidx, name_map)
            if pname is None: skipped['no_player'] += 1; continue

            # Live rolling — DNP rows already excluded by get_prior_games_played
            prior = get_prior_games_played(pidx, pname, date_str)
            if len(prior) < 5: skipped['few_games'] += 1; continue

            feats = extract_prediction_features(prior, line)
            if feats is None: skipped['no_features'] += 1; continue

            L30=feats['L30']; L20=feats['L20']; L10=feats['L10']
            L5=feats['L5'];   L3=feats['L3']
            vol=feats['vol']; trend=feats['trend']; std10=feats['std10']
            hr10=feats['hr10']; hr30=feats['hr30']
            r20=feats['recent20']; r20h=feats['r20_homes']
            fg30=feats['fg30']; fg10=feats['fg10']; fgTrend=feats['fgTrend']
            m30=feats['m30'];   m10=feats['m10'];   minTrend=feats['minTrend']
            fga30=feats['fga30']; fga10=feats['fga10']
            min_cv=feats['min_cv']; ppm=feats['ppm']; rmt=feats['rmt']; fpm=feats['fpm']

            # Team / opponent / position from most recent played game
            sn  = prior.iloc[-1]
            ta  = str(sn.get('GAME_TEAM_ABBREVIATION',''))
            ih  = ta == ht
            opp = at if ih else ht
            pos = POS_MAP.get(str(sn.get('PLAYER_POSITION','')), 'Forward')

            # H2H
            hr_  = h2h_lkp.get((pname, opp))
            hG   = int(hr_['H2H_GAMES'])        if hr_ else 0
            hA   = float(hr_['H2H_AVG_PTS'])    if hr_ else None
            hTS  = float(hr_['H2H_TS_VS_OVERALL'])   if hr_ and pd.notna(hr_.get('H2H_TS_VS_OVERALL'))   else 0
            hFA  = float(hr_['H2H_FGA_VS_OVERALL'])  if hr_ and pd.notna(hr_.get('H2H_FGA_VS_OVERALL'))  else 0
            hMN  = float(hr_['H2H_MIN_VS_OVERALL'])  if hr_ and pd.notna(hr_.get('H2H_MIN_VS_OVERALL'))  else 0
            hCF  = float(hr_['H2H_CONFIDENCE'])      if hr_ and pd.notna(hr_.get('H2H_CONFIDENCE'))      else 0
            hStr = f"{hA:.1f} ({hG}g)" if hG >= 3 and hA else ""
            uh   = hG >= 3 and hA is not None

            # Defence / pace
            dP = get_dvp(opp, pos); dO = get_def_overall(opp)
            op = pace_rank.get(opp, 15)

            # B2B
            rest = b2b.get((pname, date_str), 99); ib2b = 1 if rest == 1 else 0

            # ── Engine B: GBR projection ──
            pp = None; pg = 0
            if model:
                from model_trainer import FEATURES
                fd = {'l30':L30,'l10':L10,'l5':L5,'l3':L3,'volume':vol,'trend':trend,
                      'std10':std10,'defP':dP,'pace_rank':op,'h2h_ts_dev':hTS,
                      'h2h_fga_dev':hFA,'h2h_min_dev':hMN,'h2h_conf':hCF,
                      'min_cv':min_cv,'pts_per_min':ppm,'recent_min_trend':rmt,
                      'fga_per_min':fpm,'is_b2b':ib2b,'rest_days':rest,
                      'consistency':1/(std10+1),'line':line}
                Xp = pd.DataFrame([fd])[FEATURES].fillna(0)
                pp = float(model.predict(Xp)[0]); pg = abs(pp - line)

            # ── Engine A: 10-signal composite ──
            W = POS_WEIGHTS.get(pos, POS_WEIGHTS['Forward'])
            S = {
                1: np.clip((L30-line)/5, -1, 1),
                2: (hr30/100 - 0.5) * 2,
                3: (hr10/100 - 0.5) * 2,
                4: np.clip((L5-L30)/5, -1, 1),
                5: np.clip(vol/5, -1, 1),
                6: np.clip((dP-15)/15, -1, 1),
                7: np.clip((hA-line)/5, -1, 1) if uh else 0.0,
                8: np.clip((15-op)/15, -1, 1),
                9: np.clip((fgTrend or 0)/10, -1, 1),
               10: np.clip((minTrend or 0)/5, -1, 1),
            }
            if uh:
                tw = sum(W.values()); ws = sum(W[k]*S[k] for k in S)
            else:
                tw = sum(v for k,v in W.items() if k!=7)
                ws = sum(W[k]*S[k] for k in S if k!=7)
            comp = ws / tw if tw else 0

            # ── Direction ──
            if (pp and pp > line+0.3) or (not pp and comp > 0.05):
                dr = 'OVER'; is_lean = False
            elif (pp and pp < line-0.3) or (not pp and comp < -0.05):
                dr = 'UNDER'; is_lean = False
            else:
                raw = 'OVER' if comp >= 0 else 'UNDER'
                dr  = f'LEAN {raw}'; is_lean = True

            # ── Confidence ──
            sc = float(np.clip(0.5 + abs(comp)*0.3, 0.50, 0.85))
            if std10 > 8: sc -= 0.03
            sc   = float(np.clip(sc, 0.45, 0.85))
            pc   = float(np.clip(0.5 + pg*0.04, 0.45, 0.90)) if pp else sc
            conf = 0.4*sc + 0.6*pc

            # ── Flags ──
            io = 'UNDER' not in dr
            fl = 0; fds = []
            for nm, ag, dt in [
                ('Volume',   (io and vol>0) or (not io and vol<0),                  f"{vol:+.1f}"),
                ('HR L30',   (io and hr30>50) or (not io and hr30<50),              f"{hr30}%"),
                ('HR L10',   (io and hr10>50) or (not io and hr10<50),              f"{hr10}%"),
                ('Trend',    (io and trend>0) or (not io and trend<0),              f"{trend:+.1f}"),
                ('Context',  (io and vol>-1) or (not io and vol<1),                 f"vol={vol:+.1f}"),
                ('Defense',  (io and dP>15) or (not io and dP<15),                 f"#{dP}"),
                ('H2H',      uh and ((io and hA>line) or (not io and hA<line)),     f"{hA:.1f}" if uh else "N/A"),
                ('Pace',     (io and op<15) or (not io and op>15),                  f"#{op}"),
                ('FG Trend', fgTrend is not None and ((io and fgTrend>0) or (not io and fgTrend<0)), f"{fgTrend:+.1f}%" if fgTrend else "N/A"),
                ('Min Trend',minTrend is not None and ((io and minTrend>0) or (not io and minTrend<0)), f"{minTrend:+.1f}" if minTrend else "N/A"),
            ]:
                fl += 1 if ag else 0
                fds.append({'name':nm,'agrees':bool(ag),'detail':dt})

            # H2H alignment
            ha = True
            if hTS != 0:
                if 'OVER'  in dr and hTS < -3: ha = False
                elif 'UNDER' in dr and hTS >  3: ha = False

            # ── Tier ──
            if is_lean:
                tier=3; tl='T3_LEAN'
            elif conf>=0.70 and fl>=8 and std10<=6 and ha: tier=1; tl='T1_ULTRA'
            elif conf>=0.65 and fl>=7 and std10<=7 and ha: tier=1; tl='T1_PREMIUM'
            elif conf>=0.62 and fl>=7 and std10<=7 and ha: tier=1; tl='T1'
            elif conf>=0.55 and fl>=6 and std10<=8 and ha: tier=2; tl='T2'
            else:                                           tier=3; tl='T3'

            tr = trust.get(pname)
            if tr is not None and tr < 0.42 and tier == 1: tier=2; tl='T2'
            units = 3.0 if tl=='T1_ULTRA' else 2.0 if tier==1 else 1.0 if tier==2 else 0.0

            # ── Odds ──
            oo = american_to_decimal(pd_.get('over'))
            uo = american_to_decimal(pd_.get('under'))
            ro = sum(1 for r in r20 if r > line)
            ru = sum(1 for r in r20 if r <= line)

            # ── Line history ──
            lh  = [{'line':line,'batch':BATCH,'ts':batch_ts}]
            ekey = (pname, ms)
            if ekey in exist_map:
                ep   = exist_map[ekey]
                old_lh = ep.get('lineHistory',[])
                if isinstance(old_lh, list) and old_lh:
                    lh = old_lh
                    if not any(isinstance(h,dict) and h.get('batch')==BATCH for h in lh):
                        lh.append({'line':line,'batch':BATCH,'ts':batch_ts})
                    else:
                        for h in lh:
                            if isinstance(h,dict) and h.get('batch')==BATCH:
                                h['line']=line; h['ts']=batch_ts

            # ── Pre-match reasoning ──
            play_data = {
                'player':pname,'dir':dr,'line':line,
                'l30':L30,'l10':L10,'l5':L5,'l3':L3,
                'volume':vol,'trend':trend,'std10':std10,
                'flags':fl,'flagDetails':fds,
                'h2h':hStr,'h2hG':hG,'h2hTsDev':hTS,'h2hFgaDev':hFA,
                'h2hProfile':hr_.get('H2H_SCORING_PROFILE','') if hr_ else '',
                'defP':dP,'defO':dO,'pace':op,
                'fgTrend':fgTrend,'minTrend':minTrend,
                'minL30':m30,'minL10':m10,'conf':conf,
                'predPts':round(pp,1) if pp else None,
                'predGap':round(pg,1) if pp else None,
                'tierLabel':tl,'position':pos,'match':ms,'isHome':ih,
                'recent':r20[:5],'hr30':hr30,'hr10':hr10,
            }
            reason = generate_pre_match_reason(play_data)

            play = {
                'date':date_str,'player':pname,'match':ms,'fullMatch':fms,
                'isHome':ih,'team':ta,'gameTime':g['gt'],'position':pos,'posSimple':pos[:1],
                'line':line,'overOdds':oo,'underOdds':uo,
                'books':pd_.get('books',1),'spread':sv,'total':tv,'blowout':blow,
                'l30':round(L30,1),'l20':round(L20,1),'l10':round(L10,1),
                'l5':round(L5,1),'l3':round(L3,1),
                'hr30':hr30,'hr10':hr10,
                'recent':r20[:5],'recent10':r20[:10],'recent20':r20,'recent20homes':[bool(x) for x in r20h],
                'defO':dO,'defP':dP,'pace':op,
                'h2h':hStr,'h2hG':hG,'h2hTsDev':hTS,'h2hFgaDev':hFA,
                'h2hConfidence':hCF,'h2hProfile':hr_.get('H2H_SCORING_PROFILE','') if hr_ else '',
                'fgL30':fg30,'fgL10':fg10,'fga30':fga30,'fga10':fga10,'fg3L30':None,'fg3L10':None,
                'minL30':m30,'minL10':m10,'std10':round(std10,1),
                'dir':dr,'rawDir':dr,'conf':round(conf,3),
                'tier':tier,'tierLabel':tl,'units':units,'avail':'OK',
                'volume':vol,'trend':trend,'fgTrend':fgTrend,'minTrend':minTrend,
                'flags':fl,'flagsStr':f"{fl}/10",'flagDetails':fds,
                'recentOver':ro,'recentUnder':ru,'lineHistory':lh,
                'predPts':round(pp,1) if pp else None,'predGap':round(pg,1) if pp else None,
                'preMatchReason':reason,
                'actualPts':None,'result':None,'delta':None,
                'postMatchReason':'','lossType':None,'reason':'',
                'season':'2025-26',
            }
            plays.append(play)

    total_skip = sum(skipped.values())
    leans = sum(1 for p in plays if 'LEAN' in p.get('dir',''))
    print(f"  {len(plays)} predictions ({len(plays)-leans} conviction + {leans} leans, {total_skip} skipped)")
    if total_skip:
        print(f"    Skips: {', '.join(f'{v} {k}' for k,v in skipped.items() if v)}")
    log_event(f'B{BATCH}','PREDICTIONS',detail=f'{len(plays)} plays, skipped {total_skip}')
    return plays


# ─── SAVE TODAY.JSON ─────────────────────────────────────────
def save_today(plays, date_str):
    """
    Merge predictions into today.json.
    - Graded plays (WIN/LOSS/DNP) are IMMUTABLE
    - Ungraded plays updated with latest prediction + line history
    - Players from previous batch but not in current fetch: PRESERVED
    """
    batch_ts = now_uk().strftime('%H:%M')
    existing = []
    if TODAY_JSON.exists():
        with open(TODAY_JSON) as f: existing = json.load(f)
    before = len(existing)

    today_ex   = [p for p in existing if p.get('date') == date_str]
    historical = [p for p in existing if p.get('date') != date_str]
    ex_map     = {(p['player'], p.get('match','')): p for p in today_ex}
    new_map    = {(p['player'], p['match']): p for p in plays}

    merged = []; added = updated = preserved = 0
    for key in set(ex_map) | set(new_map):
        old = ex_map.get(key)
        new = new_map.get(key)
        if old and old.get('result') in ('WIN','LOSS','DNP'):
            merged.append(old); continue
        if old and new:
            old_lh = old.get('lineHistory',[])
            if isinstance(old_lh,list) and old_lh:
                new['lineHistory'] = old_lh
                if not any(isinstance(h,dict) and h.get('batch')==BATCH for h in old_lh):
                    new['lineHistory'].append({'line':new['line'],'batch':BATCH,'ts':batch_ts})
                else:
                    for h in new['lineHistory']:
                        if isinstance(h,dict) and h.get('batch')==BATCH:
                            h['line']=new['line']; h['ts']=batch_ts
            merged.append(new); updated += 1
        elif old and not new:
            merged.append(old); preserved += 1
        elif new and not old:
            merged.append(new); added += 1

    merged.sort(key=lambda p: (p.get('tier',9), -p.get('conf',0)))
    all_p = merged + sorted(historical, key=lambda p: p.get('date',''), reverse=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(TODAY_JSON,'w') as f: json.dump(clean_json(all_p), f)

    t1 = sum(1 for p in merged if p.get('tier')==1)
    t2 = sum(1 for p in merged if p.get('tier')==2)
    print(f"\n  ✓ today.json: {len(merged)} plays ({t1} T1, {t2} T2)")
    print(f"    Added:{added} Updated:{updated} Preserved:{preserved}")
    log_batch_summary(f'B{BATCH}', props_fetched=len(plays), plays_added=added)
    verify_no_deletion(f'B{BATCH}', TODAY_JSON, before, len(all_p), 'SAVE_TODAY')


# ─── SAVE DAILY EXCEL ────────────────────────────────────────
def save_daily_excel(plays, date_str):
    DAILY_DIR.mkdir(parents=True, exist_ok=True)
    path = DAILY_DIR / f"{date_str}.xlsx"
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        wb  = openpyxl.Workbook()
        ws1 = wb.active; ws1.title = "Props"
        ws1.append(['Player','Match','Position','Line','L3','L5','L10','L20','L30',
                    'Min_L10','Min_L30','FG_L10','FG_L30','Std10','HR_L10','HR_L30',
                    'Vol','Trend','FGTrend','MinTrend','DefP','Pace','H2H','H2H_TS_Dev'])
        for p in plays:
            ws1.append([p['player'],p['match'],p['position'],p['line'],
                        p.get('l3'),p.get('l5'),p.get('l10'),p.get('l20'),p.get('l30'),
                        p.get('minL10'),p.get('minL30'),p.get('fgL10'),p.get('fgL30'),
                        p.get('std10'),p.get('hr10'),p.get('hr30'),
                        p.get('volume'),p.get('trend'),p.get('fgTrend'),p.get('minTrend'),
                        p.get('defP'),p.get('pace'),p.get('h2h'),p.get('h2hTsDev')])
        ws2 = wb.create_sheet("Predictions")
        ws2.append(['Player','Match','Dir','TierLabel','Conf','Flags','Units',
                    'PredPts','PredGap','OverOdds','UnderOdds','Reasoning'])
        for p in plays:
            ws2.append([p['player'],p['match'],p['dir'],p['tierLabel'],
                        p.get('conf'),p.get('flags'),p.get('units'),
                        p.get('predPts'),p.get('predGap'),p.get('overOdds'),p.get('underOdds'),
                        p.get('preMatchReason','')])
        for row in ws2.iter_rows(min_row=2,min_col=12,max_col=12):
            for cell in row: cell.alignment = Alignment(wrap_text=True)
        ws2.column_dimensions[get_column_letter(12)].width = 80
        ws3 = wb.create_sheet("Graded")
        ws3.append(['Player','Match','Dir','Line','ActualPts','Result','Delta',
                    'PostMatchReason','LossType','GradedAt'])
        for p in plays:
            ws3.append([p['player'],p['match'],p['dir'],p['line'],
                        None,None,None,None,None,None])
        hf = PatternFill('solid',fgColor='1E1E2E')
        hfont = Font(bold=True,color='E0E0E0')
        for ws in [ws1,ws2,ws3]:
            for cell in ws[1]:
                cell.fill=hf; cell.font=hfont
                cell.alignment=Alignment(horizontal='center')
        wb.save(path)
        print(f"  ✓ Daily Excel: {path.name}")
    except Exception as e:
        print(f"  ⚠ Daily Excel failed: {e}")


# ─── MAIN ────────────────────────────────────────────────────
def main():
    date_str = today_et()
    if len(sys.argv) > 2 and '-' in sys.argv[2]: date_str = sys.argv[2]
    print("="*60)
    print(f"PropEdge V9.2 — BATCH {BATCH}: PREDICT")
    print(f"  Date: {date_str} | {now_uk().strftime('%Y-%m-%d %H:%M %Z')}")
    print("="*60)
    log_event(f'B{BATCH}','BATCH_START',detail=date_str)

    games, _ = fetch_props(date_str)
    if not games: print("  No games today."); return

    plays = run_predictions(games, date_str)
    save_today(plays, date_str)
    save_daily_excel(plays, date_str)

    repo = REPO_DIR if REPO_DIR.exists() else ROOT
    from batch0_grade import git_push
    git_push(repo, f"B{BATCH}: {date_str} — {len(plays)} plays")
    log_event(f'B{BATCH}','BATCH_COMPLETE')
    try:
        import subprocess
        subprocess.run(['osascript','-e',f'display notification "B{BATCH}: {len(plays)} plays" with title "PropEdge V9.2"'],capture_output=True,timeout=5)
    except: pass

if __name__ == '__main__': main()
